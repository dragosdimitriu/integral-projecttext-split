
import argparse
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import numbers

def validate_column_data(sheet, column):
    """
    Validate that:
    1. Column exists in the sheet
    2. Data exists ONLY in the specified column
    3. Data exists in the specified column
    
    Returns: (is_valid, error_message)
    """
    # Get the column index
    try:
        col_idx = column_index_from_string(column)
    except ValueError:
        return False, f"Invalid column name: '{column}'"
    
    # Check if sheet has any data
    if not sheet.dimensions:
        return False, "The sheet is empty (no data found)."
    
    # Get all columns that have data
    min_col = sheet.min_column
    max_col = sheet.max_column
    min_row = sheet.min_row
    max_row = sheet.max_row
    
    # Check if specified column is within the data range
    if col_idx < min_col or col_idx > max_col:
        return False, f"Column '{column}' does not exist in the sheet. Available columns: {get_column_letter(min_col)} to {get_column_letter(max_col)}"
    
    # Check if there's any data in the specified column
    column_has_data = False
    for row in range(min_row, max_row + 1):
        cell = sheet.cell(row=row, column=col_idx)
        if cell.value is not None and str(cell.value).strip():
            column_has_data = True
            break
    
    if not column_has_data:
        return False, f"No data found in column '{column}'. The column exists but is empty."
    
    # Check if data exists in other columns (only check columns with data)
    other_columns_with_data = []
    for col in range(min_col, max_col + 1):
        if col == col_idx:
            continue  # Skip the specified column
        
        # Check if this column has any non-empty data
        for row in range(min_row, max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip():
                col_letter = get_column_letter(col)
                other_columns_with_data.append(col_letter)
                break  # Found data in this column, move to next column
    
    if other_columns_with_data:
        return False, f"Data exists in multiple columns. Found data in columns: {', '.join(other_columns_with_data)}. Data must exist ONLY in column '{column}'."
    
    return True, None

def main(file_name, column, max_chars):
    try:
        # Open the Excel file and select the active sheet
        wb = openpyxl.load_workbook(file_name)
        
        # Validate that only ONE sheet exists
        sheet_count = len(wb.sheetnames)
        if sheet_count != 1:
            return False, f"The Excel file must contain exactly ONE sheet. Found {sheet_count} sheet(s): {', '.join(wb.sheetnames)}", None
        
        sheet = wb.active

        # Validate column data
        is_valid, error_message = validate_column_data(sheet, column)
        if not is_valid:
            return False, error_message, None

        # Select the specified column and set the number format to "text"
        column_cells = sheet[column]
        for cell in column_cells:
            cell.number_format = numbers.FORMAT_TEXT

        # Iterate through each cell in the column
        for cell in column_cells:
            # Check if the cell contains a string value and its length is > max_chars
            while isinstance(cell.value, str) and len(cell.value.strip()) > max_chars:
                # Split the cell value on the last space character that occurs before the max_chars-th character
                split_index = cell.value.rfind(' ', 0, max_chars)
                if split_index != -1:
                    # Get the column letter of the new column
                    new_column = get_column_letter(cell.column + 1)
                    # Insert a new column to the right of the current column
                    #sheet.insert_cols(cell.column + 1, amount=1, idx=cell.column + 1)
                    # Move the remaining text to the new column
                    sheet.cell(row=cell.row, column=cell.column + 1).value = cell.value[split_index+1:].strip()
                    # Truncate the original cell value to the first part
                    cell.value = cell.value[:split_index].strip()
                    # Set the number format of the new column to "text"
                    #sheet[new_column][cell.row].number_format = numbers.FORMAT_TEXT
                    # Check if the new cell also needs to be split
                    cell = sheet.cell(row=cell.row, column=cell.column + 1)
                else:
                    # Cannot split the cell any further, move to the next cell
                    break
         # Save the modified Excel file
        output_filename = f"{file_name.split('.')[0]}_ProjectTextReady.xlsx"
        wb.save(output_filename)
        return True, "File successfully processed.", output_filename

    except Exception as e:
        return False, str(e), None


if __name__ == '__main__':
    # Create argument parser
    parser = argparse.ArgumentParser(description='Split cells in a given column if they exceed a maximum length.')
    parser.add_argument('file_name', type=str, help='the name of the Excel file to be processed')
    parser.add_argument('column', type=str, help='the column to be processed')
    parser.add_argument('max_chars', type=int, help='the maximum number of characters allowed in a cell')

    # Parse command line arguments
    args = parser.parse_args()

    # Call main function with command line arguments
    main(args.file_name, args.column, args.max_chars)