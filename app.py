import os
import sys
import time
import re
import struct
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from authlib.integrations.flask_client import OAuth
from flask_mail import Mail, Message
from dotenv import load_dotenv
from datetime import datetime, timedelta
import split
import uuid
import openpyxl
import json
from collections import defaultdict

# Force unbuffered output for better debugging
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except:
    # Fallback for older Python versions
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, line_buffering=True)

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Session configuration
# Set SESSION_COOKIE_SECURE=True in production when using HTTPS
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', 'False').lower() == 'true'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Shorter session timeout: 2 hours (was 7 days)
app.permanent_session_lifetime = timedelta(hours=2)

# Google OAuth Configuration
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', '')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', '')
app.config['GOOGLE_DISCOVERY_URL'] = "https://accounts.google.com/.well-known/openid-configuration"

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'False').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', '')
app.config['MAIL_ENABLED'] = os.environ.get('MAIL_ENABLED', 'False').lower() == 'true'

# Initialize Flask-Mail
mail = Mail(app)

# Debug: Print email configuration on startup
if app.config['MAIL_ENABLED']:
    print("\n" + "="*60)
    print("EMAIL CONFIGURATION (on startup):")
    print("="*60)
    print(f"MAIL_ENABLED: {app.config['MAIL_ENABLED']}")
    print(f"MAIL_SERVER: {app.config['MAIL_SERVER']}")
    print(f"MAIL_PORT: {app.config['MAIL_PORT']}")
    print(f"MAIL_USE_TLS: {app.config['MAIL_USE_TLS']}")
    print(f"MAIL_USE_SSL: {app.config['MAIL_USE_SSL']}")
    print(f"MAIL_USERNAME: {app.config['MAIL_USERNAME']}")
    print(f"MAIL_PASSWORD: {'*' * len(app.config['MAIL_PASSWORD']) if app.config['MAIL_PASSWORD'] else 'NOT SET'}")
    print(f"MAIL_DEFAULT_SENDER: {app.config['MAIL_DEFAULT_SENDER']}")
    print("="*60 + "\n")
else:
    print("EMAIL DEBUG: Email notifications are DISABLED (MAIL_ENABLED=False)")

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# Initialize OAuth
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=app.config['GOOGLE_CLIENT_ID'],
    client_secret=app.config['GOOGLE_CLIENT_SECRET'],
    server_metadata_url=app.config['GOOGLE_DISCOVERY_URL'],
    client_kwargs={
        'scope': 'openid email profile',
        'prompt': 'select_account'
    }
)

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, email, name, picture):
        self.id = id
        self.email = email
        self.name = name
        self.picture = picture

@login_manager.user_loader
def load_user(user_id):
    # Load user from session
    if 'user_info' in session:
        user_info = session['user_info']
        # Verify the user_id matches
        if user_info.get('id') == user_id:
            return User(
                id=user_info['id'],
                email=user_info['email'],
                name=user_info['name'],
                picture=user_info.get('picture', '')
            )
    return None

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Statistics tracking file
STATS_FILE = 'processing_stats.json'

def load_stats():
    """Load processing statistics from file"""
    if os.path.exists(STATS_FILE):
        try:
            with open(STATS_FILE, 'r') as f:
                return json.load(f)
        except:
            pass
    return {
        'total_processed': 0,
        'total_successful': 0,
        'total_failed': 0,
        'total_processing_time': 0.0,
        'processing_history': []
    }

def save_stats(stats):
    """Save processing statistics to file"""
    try:
        with open(STATS_FILE, 'w') as f:
            json.dump(stats, f, indent=2)
    except Exception as e:
        print(f"Error saving stats: {e}")

def add_processing_record(success, processing_time, user_email=None):
    """Add a processing record to statistics"""
    stats = load_stats()
    stats['total_processed'] += 1
    if success:
        stats['total_successful'] += 1
    else:
        stats['total_failed'] += 1
    
    if processing_time:
        stats['total_processing_time'] += processing_time
    
    # Add to history (keep last 1000 records)
    stats['processing_history'].append({
        'timestamp': datetime.now().isoformat(),
        'success': success,
        'processing_time': processing_time,
        'user_email': user_email
    })
    
    # Keep only last 1000 records
    if len(stats['processing_history']) > 1000:
        stats['processing_history'] = stats['processing_history'][-1000:]
    
    save_stats(stats)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_COLUMN_LENGTH = 3  # Maximum column name length (e.g., "ZZZ")
MAX_CHARS_LIMIT = 23  # Maximum characters per cell limit (recommended: 18-20)
MIN_CHARS_LIMIT = 18  # Minimum characters per cell limit
SUGGESTED_MAX_CHARS = 20  # Suggested optimal value

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_column_name(column):
    """Validate column name format (only letters, max 3 characters)"""
    if not column:
        return False, "Column name cannot be empty"
    if len(column) > MAX_COLUMN_LENGTH:
        return False, f"Column name must be {MAX_COLUMN_LENGTH} characters or less"
    if not re.match(r'^[A-Z]+$', column):
        return False, "Column name must contain only uppercase letters (A-Z)"
    return True, None

def validate_max_chars(max_chars):
    """Validate max_chars is within reasonable bounds (18-23, recommended 18-20)"""
    if max_chars < MIN_CHARS_LIMIT:
        return False, f"Max characters must be at least {MIN_CHARS_LIMIT} (recommended: 18-20)"
    if max_chars > MAX_CHARS_LIMIT:
        return False, f"Max characters cannot exceed {MAX_CHARS_LIMIT} (recommended: 18-20)"
    return True, None

def validate_excel_file(filepath):
    """Validate that the file is a valid Excel file"""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        wb.close()
        return True, None
    except Exception as e:
        return False, f"Invalid Excel file: {str(e)}"

def validate_file_content(filepath):
    """Enhanced file validation: Check file signature/MIME type"""
    try:
        with open(filepath, 'rb') as f:
            # Read first 8 bytes to check file signature
            header = f.read(8)
            
            # Excel file signatures:
            # .xlsx: PK\x03\x04 (ZIP archive, Excel 2007+)
            # .xls: \xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1 (OLE2, Excel 97-2003)
            
            if len(header) < 8:
                return False, "File too small or corrupted"
            
            # Check for .xlsx (ZIP-based format)
            if header[:2] == b'PK':
                # Verify it's actually a ZIP file
                if header[2:4] == b'\x03\x04':
                    return True, None
            
            # Check for .xls (OLE2 format)
            if header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
                return True, None
            
            return False, "File does not appear to be a valid Excel file (invalid file signature)"
    except Exception as e:
        return False, f"Error reading file: {str(e)}"

def sanitize_filename(filename):
    """Additional sanitization for filename"""
    # Remove any path components
    filename = os.path.basename(filename)
    # Remove any dangerous characters (keep only alphanumeric, dots, underscores, hyphens)
    filename = re.sub(r'[^a-zA-Z0-9._-]', '', filename)
    # Limit filename length
    if len(filename) > 255:
        name, ext = os.path.splitext(filename)
        filename = name[:255-len(ext)] + ext
    return filename

def sanitize_input(text, max_length=100):
    """Sanitize and validate text input"""
    if text is None:
        return None
    # Convert to string if it's not already
    if not isinstance(text, str):
        text = str(text)
    # Remove leading/trailing whitespace
    text = text.strip()
    # Limit length
    if len(text) > max_length:
        text = text[:max_length]
    # Remove any control characters
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    return text

def validate_email(email):
    """Validate email format"""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))

@app.route('/')
@login_required
def index():
    # Ensure user is properly loaded
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('index.html', user=current_user)

@app.route('/login')
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/auth/google')
def google_login():
    redirect_uri = url_for('callback', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/callback')
def callback():
    try:
        token = google.authorize_access_token()
        
        if not token:
            return render_template('login.html', error="Failed to get access token from Google"), 400
        
        # Fetch user info from Google - Authlib handles the userinfo endpoint automatically
        # Try using the token directly with the userinfo endpoint
        resp = google.get('https://www.googleapis.com/oauth2/v2/userinfo', token=token)
        
        if resp.status_code != 200:
            # Try alternative endpoint
            resp = google.get('https://openidconnect.googleapis.com/v1/userinfo', token=token)
            if resp.status_code != 200:
                return render_template('login.html', error=f"Failed to fetch user info: HTTP {resp.status_code}"), 400
        
        user_info = resp.json()
        
        # Create user object - handle both 'sub' and 'id' fields
        user_id = user_info.get('sub') or user_info.get('id', '')
        if not user_id:
            return render_template('login.html', error="Failed to get user ID from Google"), 400
        
        user_email = user_info.get('email', '')
        user_name = user_info.get('name', user_email or 'Unknown')
        user_picture = user_info.get('picture', '')
        
        # Create user object
        user = User(
            id=user_id,
            email=user_email,
            name=user_name,
            picture=user_picture
        )
        
        # Store user info in session BEFORE login_user
        session['user_info'] = {
            'id': user.id,
            'email': user.email,
            'name': user.name,
            'picture': user.picture
        }
        
        # Make session permanent (2 hour timeout)
        session.permanent = True
        
        # Login the user (don't use remember=True to enforce session timeout)
        login_user(user, remember=False)
        
        return redirect(url_for('index'))
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Authentication error: {error_details}")  # Debug output
        return render_template('login.html', error=f"Error during authentication: {str(e)}"), 400

@app.route('/logout')
@login_required
def logout():
    # Properly invalidate session
    user_id = current_user.id if current_user.is_authenticated else None
    logout_user()
    session.clear()
    # Regenerate session ID to prevent session fixation
    session.permanent = False
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Validate file extension
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls files'}), 400
    
    # Sanitize filename
    original_filename = sanitize_filename(secure_filename(file.filename))
    if not original_filename:
        return jsonify({'success': False, 'error': 'Invalid filename'}), 400
    
    # Generate unique filename to avoid conflicts
    file_id = str(uuid.uuid4())
    filename = f"{file_id}_{original_filename}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # Ensure path is within upload folder (prevent path traversal)
    filepath = os.path.normpath(filepath)
    if not filepath.startswith(os.path.normpath(app.config['UPLOAD_FOLDER'])):
        return jsonify({'success': False, 'error': 'Invalid file path'}), 400
    
    try:
        file.save(filepath)
        
        # Enhanced file validation: Check file content/signature
        is_valid, error_msg = validate_file_content(filepath)
        if not is_valid:
            # Clean up invalid file
            try:
                os.remove(filepath)
            except:
                pass
            return jsonify({'success': False, 'error': error_msg}), 400
        
        # Validate that the file is actually a valid Excel file (can be opened)
        is_valid, error_msg = validate_excel_file(filepath)
        if not is_valid:
            # Clean up invalid file
            try:
                os.remove(filepath)
            except:
                pass
            return jsonify({'success': False, 'error': error_msg}), 400
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error saving file: {str(e)}'}), 500
    
    return jsonify({
        'success': True,
        'file_id': file_id,
        'filename': original_filename,
        'uploaded_filename': filename,
        'upload_time': datetime.now().isoformat()
    })

@app.route('/process', methods=['POST'])
@login_required
def process_file():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Invalid request data'}), 400
    
    # Enhanced input sanitization
    uploaded_filename = sanitize_input(data.get('uploaded_filename'), max_length=300)
    column = sanitize_input(data.get('column'), max_length=3)
    if column:
        column = column.upper()
    
    # Handle max_chars - it might be int or string from JSON
    max_chars_value = data.get('max_chars')
    if max_chars_value is None:
        return jsonify({'success': False, 'error': 'Missing max_chars parameter'}), 400
    
    if not all([uploaded_filename, column]):
        return jsonify({'success': False, 'error': 'Missing required parameters'}), 400
    
    # Validate column name again
    is_valid, error_msg = validate_column_name(column)
    if not is_valid:
        return jsonify({'success': False, 'error': error_msg}), 400
    
    # Validate max_chars - convert to int (handles both int and string)
    try:
        max_chars = int(max_chars_value)
        is_valid, error_msg = validate_max_chars(max_chars)
        if not is_valid:
            return jsonify({'success': False, 'error': error_msg}), 400
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'Invalid max characters value'}), 400
    
    # Sanitize filename to prevent path traversal
    uploaded_filename = sanitize_filename(uploaded_filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], uploaded_filename)
    
    # Ensure path is within upload folder (prevent path traversal)
    filepath = os.path.normpath(filepath)
    if not filepath.startswith(os.path.normpath(app.config['UPLOAD_FOLDER'])):
        return jsonify({'success': False, 'error': 'Invalid file path'}), 400
    
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    try:
        # Start timing
        start_time = time.time()
        
        # Process the file using split.py
        success, message, output_filename = split.main(filepath, column, max_chars)
        
        # Calculate processing time
        processing_time = time.time() - start_time
        
        if success:
            # Move output file to outputs folder
            if output_filename and os.path.exists(output_filename):
                output_basename = os.path.basename(output_filename)
                output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_basename)
                os.rename(output_filename, output_path)
                
                # Send email notification if enabled
                if app.config['MAIL_ENABLED'] and current_user.is_authenticated:
                    print(f"\nEMAIL DEBUG: Attempting to send email to {current_user.email}")
                    try:
                        send_processing_complete_email(
                            current_user.email,
                            current_user.name,
                            uploaded_filename,
                            output_basename,
                            round(processing_time, 2)
                        )
                    except Exception as e:
                        # Don't fail the request if email fails
                        print(f"EMAIL DEBUG: Exception caught in process_file: {str(e)}")
                        import traceback
                        traceback.print_exc()
                elif not app.config['MAIL_ENABLED']:
                    print("EMAIL DEBUG: Email disabled or user not authenticated - skipping email")
                elif not current_user.is_authenticated:
                    print("EMAIL DEBUG: User not authenticated - skipping email")
                
                # Track successful processing
                user_email = current_user.email if current_user.is_authenticated else None
                add_processing_record(True, round(processing_time, 2), user_email)
                
                return jsonify({
                    'success': True,
                    'message': message,
                    'output_filename': output_basename,
                    'processing_time': round(processing_time, 2)
                })
            else:
                # Track failed processing
                user_email = current_user.email if current_user.is_authenticated else None
                add_processing_record(False, round(processing_time, 2), user_email)
                return jsonify({'success': False, 'error': 'Output file was not created'}), 500
        else:
            # Track failed processing
            user_email = current_user.email if current_user.is_authenticated else None
            add_processing_record(False, 0, user_email)
            return jsonify({'success': False, 'error': message}), 500
            
    except Exception as e:
        # Track failed processing
        user_email = current_user.email if current_user.is_authenticated else None
        add_processing_record(False, 0, user_email)
        return jsonify({'success': False, 'error': str(e)}), 500

def send_processing_complete_email(user_email, user_name, input_filename, output_filename, processing_time):
    """Send email notification when processing completes"""
    print("\n" + "="*60)
    print("EMAIL DEBUG: Starting email send process")
    print("="*60)
    
    if not app.config['MAIL_ENABLED']:
        print("EMAIL DEBUG: MAIL_ENABLED is False - email disabled")
        print("="*60 + "\n")
        return
    
    print(f"EMAIL DEBUG: MAIL_ENABLED = {app.config['MAIL_ENABLED']}")
    print(f"EMAIL DEBUG: MAIL_SERVER = {app.config['MAIL_SERVER']}")
    print(f"EMAIL DEBUG: MAIL_PORT = {app.config['MAIL_PORT']}")
    print(f"EMAIL DEBUG: MAIL_USE_TLS = {app.config['MAIL_USE_TLS']}")
    print(f"EMAIL DEBUG: MAIL_USE_SSL = {app.config['MAIL_USE_SSL']}")
    print(f"EMAIL DEBUG: MAIL_USERNAME = {app.config['MAIL_USERNAME']}")
    print(f"EMAIL DEBUG: MAIL_PASSWORD = {'*' * len(app.config['MAIL_PASSWORD']) if app.config['MAIL_PASSWORD'] else 'NOT SET'}")
    print(f"EMAIL DEBUG: MAIL_DEFAULT_SENDER = {app.config['MAIL_DEFAULT_SENDER']}")
    
    try:
        # Get base URL from environment or use default
        base_url = os.environ.get('BASE_URL', '')
        if not base_url:
            try:
                base_url = request.host_url.rstrip('/')
                print(f"EMAIL DEBUG: Using request host URL: {base_url}")
            except:
                base_url = 'https://pt.schrack.lastchance.ro'
                print(f"EMAIL DEBUG: Using default URL: {base_url}")
        else:
            print(f"EMAIL DEBUG: Using BASE_URL from env: {base_url}")
        
        download_url = f"{base_url}/download/outputs/{output_filename}"
        print(f"EMAIL DEBUG: Download URL = {download_url}")
        
        subject = f"Procesare Completă: {input_filename}"
        
        # Create beautiful HTML email optimized for Gmail
        html_body = f"""
        <!DOCTYPE html>
        <html lang="ro">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #f5f7fa;">
            <table role="presentation" style="width: 100%; border-collapse: collapse; background-color: #f5f7fa; padding: 40px 20px;">
                <tr>
                    <td align="center">
                        <table role="presentation" style="max-width: 600px; width: 100%; background-color: #ffffff; border-radius: 12px; box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1); overflow: hidden;">
                            <!-- Header -->
                            <tr>
                                <td style="background: linear-gradient(135deg, #4A478A 0%, #3A356A 100%); padding: 40px 30px; text-align: center;">
                                    <h1 style="margin: 0; color: #ffffff; font-size: 24px; font-weight: 600;">Integral ProjectText FileProcessor</h1>
                                    <p style="margin: 8px 0 0 0; color: rgba(255, 255, 255, 0.9); font-size: 14px;">Procesare Completă cu Succes</p>
                                </td>
                            </tr>
                            
                            <!-- Content -->
                            <tr>
                                <td style="padding: 40px 30px;">
                                    <p style="margin: 0 0 20px 0; color: #1e293b; font-size: 16px; line-height: 1.6;">
                                        Salut <strong>{user_name}</strong>,
                                    </p>
                                    
                                    <p style="margin: 0 0 30px 0; color: #64748b; font-size: 15px; line-height: 1.6;">
                                        Fișierul tău a fost procesat cu succes! Iată detaliile:
                                    </p>
                                    
                                    <!-- File Details Card -->
                                    <table role="presentation" style="width: 100%; border-collapse: collapse; background-color: #f8fafc; border-radius: 8px; margin-bottom: 30px; overflow: hidden;">
                                        <tr>
                                            <td style="padding: 20px;">
                                                <table role="presentation" style="width: 100%; border-collapse: collapse;">
                                                    <tr>
                                                        <td style="padding: 8px 0; border-bottom: 1px solid #e2e8f0;">
                                                            <span style="color: #64748b; font-size: 14px; font-weight: 500;">Fișier Intrare:</span>
                                                            <span style="color: #1e293b; font-size: 14px; margin-left: 8px;">{input_filename}</span>
                                                        </td>
                                                    </tr>
                                                    <tr>
                                                        <td style="padding: 8px 0; border-bottom: 1px solid #e2e8f0;">
                                                            <span style="color: #64748b; font-size: 14px; font-weight: 500;">Fișier Ieșire:</span>
                                                            <span style="color: #1e293b; font-size: 14px; margin-left: 8px;">{output_filename}</span>
                                                        </td>
                                                    </tr>
                                                    <tr>
                                                        <td style="padding: 8px 0;">
                                                            <span style="color: #64748b; font-size: 14px; font-weight: 500;">Timp Procesare:</span>
                                                            <span style="color: #1e293b; font-size: 14px; margin-left: 8px;">{processing_time} secunde</span>
                                                        </td>
                                                    </tr>
                                                </table>
                                            </td>
                                        </tr>
                                    </table>
                                    
                                    <!-- Download Button -->
                                    <table role="presentation" style="width: 100%; border-collapse: collapse; margin-bottom: 30px;">
                                        <tr>
                                            <td align="center">
                                                <a href="{download_url}" style="display: inline-block; background: linear-gradient(135deg, #4A478A 0%, #3A356A 100%); color: #ffffff; text-decoration: none; padding: 14px 32px; border-radius: 8px; font-weight: 600; font-size: 16px; box-shadow: 0 4px 12px rgba(74, 71, 138, 0.3);">
                                                    📥 Descarcă Fișierul Procesat
                                                </a>
                                            </td>
                                        </tr>
                                    </table>
                                    
                                    <p style="margin: 0 0 10px 0; color: #64748b; font-size: 13px; line-height: 1.5; text-align: center;">
                                        Linkul va rămâne valabil timp de 7 zile.
                                    </p>
                                    
                                    <p style="margin: 20px 0 0 0; color: #64748b; font-size: 13px; line-height: 1.5; text-align: center;">
                                        Dacă butonul nu funcționează, copiază și lipește acest link în browser:<br>
                                        <a href="{download_url}" style="color: #4A478A; word-break: break-all;">{download_url}</a>
                                    </p>
                                </td>
                            </tr>
                            
                            <!-- Footer -->
                            <tr>
                                <td style="background-color: #f8fafc; padding: 20px 30px; text-align: center; border-top: 1px solid #e2e8f0;">
                                    <p style="margin: 0; color: #64748b; font-size: 12px; line-height: 1.5;">
                                        Cu respect,<br>
                                        <strong style="color: #4A478A;">Echipa Integral ProjectText FileProcessor</strong>
                                    </p>
                                    <p style="margin: 12px 0 0 0; color: #94a3b8; font-size: 11px;">
                                        Acest email a fost generat automat. Vă rugăm să nu răspundeți.
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """
        
        # Plain text fallback
        text_body = f"""Salut {user_name},

Fișierul tău a fost procesat cu succes!

Fișier Intrare: {input_filename}
Fișier Ieșire: {output_filename}
Timp Procesare: {processing_time} secunde

Descarcă fișierul procesat:
{download_url}

Linkul va rămâne valabil timp de 7 zile.

Cu respect,
Echipa Integral ProjectText FileProcessor"""
        
        # Determine sender - use MAIL_DEFAULT_SENDER or fallback to MAIL_USERNAME
        sender = app.config['MAIL_DEFAULT_SENDER']
        if not sender:
            # Fallback to MAIL_USERNAME if it's a valid email
            sender = app.config['MAIL_USERNAME'] if app.config['MAIL_USERNAME'] and '@' in app.config['MAIL_USERNAME'] else None
            print(f"EMAIL DEBUG: Sender not set, using fallback: {sender}")
        
        if not sender:
            print("EMAIL DEBUG: WARNING - MAIL_DEFAULT_SENDER not set. Email may fail.")
            sender = app.config['MAIL_USERNAME'] if app.config['MAIL_USERNAME'] else 'noreply@example.com'
            print(f"EMAIL DEBUG: Using fallback sender: {sender}")
        else:
            print(f"EMAIL DEBUG: Using sender: {sender}")
        
        print(f"EMAIL DEBUG: Recipient: {user_email}")
        print(f"EMAIL DEBUG: Subject: {subject}")
        print(f"EMAIL DEBUG: HTML body length: {len(html_body)} characters")
        
        # Create message with both HTML and plain text (multipart)
        msg = Message(
            subject=subject,
            recipients=[user_email],
            html=html_body,
            body=text_body,
            sender=sender
        )
        
        print("EMAIL DEBUG: Attempting to send email...")
        try:
            mail.send(msg)
            print("EMAIL DEBUG: ✓ Flask-Mail send() completed without exception")
            print("EMAIL DEBUG: Email sent successfully via Gmail SMTP")
            print("EMAIL DEBUG: Check spam folder if email is not received")
        except Exception as send_error:
            print(f"EMAIL DEBUG: ✗ Flask-Mail send() raised exception: {type(send_error).__name__}")
            print(f"EMAIL DEBUG: Error: {str(send_error)}")
            raise  # Re-raise to be caught by outer exception handler
        print("="*60 + "\n")
        
    except Exception as e:
        # Don't raise - email failure shouldn't break the request
        print(f"EMAIL DEBUG: ✗ ERROR sending email notification")
        print(f"EMAIL DEBUG: Error type: {type(e).__name__}")
        print(f"EMAIL DEBUG: Error message: {str(e)}")
        import traceback
        print("EMAIL DEBUG: Full traceback:")
        traceback.print_exc()
        print("="*60 + "\n")

@app.route('/api/statistics')
@login_required
def get_statistics():
    """Get processing statistics"""
    stats = load_stats()
    
    # Calculate averages
    avg_processing_time = 0.0
    if stats['total_successful'] > 0:
        avg_processing_time = stats['total_processing_time'] / stats['total_successful']
    
    success_rate = 0.0
    if stats['total_processed'] > 0:
        success_rate = (stats['total_successful'] / stats['total_processed']) * 100
    
    return jsonify({
        'success': True,
        'statistics': {
            'total_processed': stats['total_processed'],
            'total_successful': stats['total_successful'],
            'total_failed': stats['total_failed'],
            'average_processing_time': round(avg_processing_time, 2),
            'success_rate': round(success_rate, 2)
        }
    })

@app.route('/api/validate-file', methods=['POST'])
@login_required
def validate_file_advanced():
    """Advanced file validation - check if file has only one sheet and only one column with data"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    # Validate file extension
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls files'}), 400
    
    # Save file temporarily for validation
    temp_filename = f"temp_{uuid.uuid4()}_{secure_filename(file.filename)}"
    temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
    
    try:
        file.save(temp_filepath)
        
        # Validate file content/signature
        is_valid, error_msg = validate_file_content(temp_filepath)
        if not is_valid:
            try:
                os.remove(temp_filepath)
            except:
                pass
            return jsonify({'success': False, 'error': error_msg}), 400
        
        # Validate Excel file can be opened
        is_valid, error_msg = validate_excel_file(temp_filepath)
        if not is_valid:
            try:
                os.remove(temp_filepath)
            except:
                pass
            return jsonify({'success': False, 'error': error_msg}), 400
        
        # Advanced validation: Check sheets and columns
        wb = openpyxl.load_workbook(temp_filepath, read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        
        if sheet_count != 1:
            wb.close()
            try:
                os.remove(temp_filepath)
            except:
                pass
            return jsonify({
                'success': False,
                'error': f'File must contain exactly ONE sheet. Found {sheet_count} sheet(s): {", ".join(wb.sheetnames)}',
                'sheet_count': sheet_count,
                'sheet_names': wb.sheetnames
            }), 400
        
        sheet = wb.active
        
        # Check which columns have data
        columns_with_data = []
        from openpyxl.utils import get_column_letter
        
        if sheet.max_column > 0 and sheet.max_row > 0:
            for col_idx in range(1, sheet.max_column + 1):
                has_data = False
                for row_idx in range(1, min(sheet.max_row + 1, 1000)):  # Check first 1000 rows
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        has_data = True
                        break
                
                if has_data:
                    col_letter = get_column_letter(col_idx)
                    # Get sample data from first non-empty cell
                    sample_value = ''
                    for row_idx in range(1, min(sheet.max_row + 1, 100)):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and str(cell.value).strip():
                            sample_value = str(cell.value)[:50]
                            break
                    
                    # Calculate average cell length in this column
                    cell_lengths = []
                    for row_idx in range(1, min(sheet.max_row + 1, 100)):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, str):
                            cell_lengths.append(len(cell.value.strip()))
                    
                    avg_length = sum(cell_lengths) / len(cell_lengths) if cell_lengths else 0
                    max_length = max(cell_lengths) if cell_lengths else 0
                    
                    columns_with_data.append({
                        'letter': col_letter,
                        'index': col_idx,
                        'sample': sample_value,
                        'avg_length': round(avg_length, 0),
                        'max_length': max_length
                    })
        
        wb.close()
        
        # Clean up temp file
        try:
            os.remove(temp_filepath)
        except:
            pass
        
        if len(columns_with_data) == 0:
            return jsonify({
                'success': False,
                'error': 'No data found in the file. The file appears to be empty.',
                'columns_with_data': []
            }), 400
        
        if len(columns_with_data) > 1:
            column_letters = [col['letter'] for col in columns_with_data]
            return jsonify({
                'success': False,
                'error': f'Data exists in multiple columns: {", ".join(column_letters)}. Data must exist ONLY in one column.',
                'columns_with_data': columns_with_data
            }), 400
        
        # File is valid - suggest optimal parameters
        single_column = columns_with_data[0]
        # Always suggest 20 characters (optimal range is 18-20, max 23)
        suggested_max_chars = 20
        
        return jsonify({
            'success': True,
            'valid': True,
            'sheet_count': 1,
            'column_with_data': single_column['letter'],
            'suggested_parameters': {
                'column': single_column['letter'],
                'max_chars': suggested_max_chars,
                'reason': 'Optimal value: 20 characters (recommended range: 18-20, maximum: 23)'
            },
            'file_info': {
                'total_rows': sheet.max_row,
                'total_columns': sheet.max_column,
                'column_stats': single_column
            }
        })
        
    except Exception as e:
        # Clean up temp file on error
        try:
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
        except:
            pass
        return jsonify({'success': False, 'error': f'Error validating file: {str(e)}'}), 500

@app.route('/preview/<folder>/<filename>')
@login_required
def preview_file(folder, filename):
    """Preview Excel file with pagination - 50 rows per page"""
    if folder not in ['uploads', 'outputs']:
        return jsonify({'success': False, 'error': 'Invalid folder'}), 400
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    rows_per_page = 50
    
    # Sanitize filename to prevent path traversal
    filename = sanitize_filename(filename)
    filepath = os.path.join(folder, filename)
    
    # Ensure path is within allowed folder (prevent path traversal)
    filepath = os.path.normpath(filepath)
    allowed_path = os.path.normpath(folder)
    if not filepath.startswith(allowed_path):
        return jsonify({'success': False, 'error': 'Invalid file path'}), 400
    
    if not os.path.exists(filepath):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    try:
        # Load workbook in read-only mode for better performance
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        
        # Get total dimensions
        total_rows = sheet.max_row
        total_cols = sheet.max_column
        
        # Limit columns for very wide files
        preview_col_count = min(50, total_cols)
        
        # Calculate pagination
        total_pages = (total_rows + rows_per_page - 1) // rows_per_page  # Ceiling division
        if total_pages == 0:
            total_pages = 1  # At least 1 page even if no rows
        page = max(1, min(page, total_pages))  # Clamp page between 1 and total_pages
        
        # Calculate row range for this page
        start_row = (page - 1) * rows_per_page + 1
        end_row = min(start_row + rows_per_page - 1, total_rows)
        
        # Get preview data for current page
        preview_data = []
        
        # Use iter_rows for better memory efficiency
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, 
                                   min_col=1, max_col=preview_col_count, 
                                   values_only=True):
            row_data = []
            for cell_value in row:
                # Get cell value, limit string length for preview
                if cell_value is None:
                    row_data.append('')
                elif isinstance(cell_value, str):
                    # Limit string length to prevent huge JSON responses
                    if len(cell_value) > 200:
                        row_data.append(cell_value[:200] + '...')
                    else:
                        row_data.append(cell_value)
                else:
                    # Convert other types to string
                    str_value = str(cell_value)
                    if len(str_value) > 200:
                        row_data.append(str_value[:200] + '...')
                    else:
                        row_data.append(str_value)
            preview_data.append(row_data)
        
        wb.close()
        
        # Debug: Print pagination info
        print(f"DEBUG Preview: total_rows={total_rows}, rows_per_page={rows_per_page}, total_pages={total_pages}, page={page}, start_row={start_row}, end_row={end_row}, data_rows={len(preview_data)}")
        
        return jsonify({
            'success': True,
            'filename': filename,
            'sheet_name': sheet.title,
            'total_rows': total_rows,
            'total_columns': total_cols,
            'preview_columns': preview_col_count,
            'has_more_columns': total_cols > 50,
            'current_page': page,
            'total_pages': total_pages,
            'rows_per_page': rows_per_page,
            'start_row': start_row,
            'end_row': end_row,
            'data': preview_data
        })
    except MemoryError:
        return jsonify({'success': False, 'error': 'Fișierul este prea mare pentru previzualizare. Vă rugăm să-l descărcați pentru a-l vedea complet.'}), 413
    except Exception as e:
        # Ensure we always return JSON, not HTML
        error_msg = str(e)
        if len(error_msg) > 200:
            error_msg = error_msg[:200] + '...'
        return jsonify({'success': False, 'error': f'Eroare la citirea fișierului: {error_msg}'}), 500

@app.route('/download/<folder>/<filename>')
@login_required
def download_file(folder, filename):
    if folder not in ['uploads', 'outputs']:
        return jsonify({'error': 'Invalid folder'}), 400
    
    # Sanitize filename to prevent path traversal
    filename = sanitize_filename(filename)
    filepath = os.path.join(folder, filename)
    
    # Ensure path is within allowed folder (prevent path traversal)
    filepath = os.path.normpath(filepath)
    allowed_path = os.path.normpath(folder)
    if not filepath.startswith(allowed_path):
        return jsonify({'error': 'Invalid file path'}), 400
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(filepath, as_attachment=True)

# Add request logging middleware
@app.before_request
def log_request_info():
    """Log all incoming requests"""
    print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {request.method} {request.path}")
    if request.method == 'POST':
        print(f"  Content-Type: {request.content_type}")
        if request.is_json:
            print(f"  JSON Data: {request.json}")
    sys.stdout.flush()

@app.after_request
def log_response_info(response):
    """Log all responses"""
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {response.status_code} {request.method} {request.path}")
    sys.stdout.flush()
    return response

if __name__ == '__main__':
    # Development server only - DO NOT use in production!
    # Use gunicorn, waitress, or another WSGI server for production
    # See DEPLOYMENT.md for production deployment instructions
    flask_debug_raw = os.environ.get('FLASK_DEBUG', 'True')
    debug_mode = flask_debug_raw.lower() == 'true'  # Default to True for local dev
    print(f"DEBUG: FLASK_DEBUG from environment: '{flask_debug_raw}' -> debug_mode={debug_mode}", flush=True)
    
    # Enable logging for better visibility
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        force=True  # Force reconfiguration
    )
    
    # Enable Werkzeug request logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.INFO)
    log.disabled = False  # Make sure it's enabled
    
    # Print startup message
    print("\n" + "="*60, flush=True)
    print("Starting Flask Development Server", flush=True)
    print("="*60, flush=True)
    print(f"Debug mode: {debug_mode}", flush=True)
    print(f"Host: 0.0.0.0", flush=True)
    print(f"Port: 5001", flush=True)
    print(f"URL: http://localhost:5001", flush=True)
    print("="*60, flush=True)
    print("\nRequest logs will appear below...", flush=True)
    print("-" * 60 + "\n", flush=True)
    
    # Use use_reloader=False if output is not showing (reloader can suppress output)
    use_reloader = debug_mode and os.environ.get('FLASK_USE_RELOADER', 'True').lower() == 'true'
    
    # Force Python to use unbuffered mode
    os.environ['PYTHONUNBUFFERED'] = '1'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=5001, use_reloader=use_reloader)

